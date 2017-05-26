# coding: utf-8
from openpyxl import load_workbook
import os
import simplejson as json

path = os.path.dirname(os.path.abspath(__file__))

# Columns of origin countries.
origins = [9, 240]

# Column of destination country names
country_names = "B"

# Columns for each year and gender of total migration populaton
total_pops = {
    '1990': {
        "t": "F",
        "m": "L",
        "f": "R"
    },
    '1995': {
        "t": "G",
        "m": "M",
        "f": "S"
    },
    '2000': {
        "t": "H",
        "m": "N",
        "f": "T"
    },
    '2005': {
        "t": "I",
        "m": "O",
        "f": "U"
    },
    '2010': {
        "t": "J",
        "m": "P",
        "f": "V"
    },
    '2015': {
        "t": "K",
        "m": "Q",
        "f": "W"
    }
}

# Rows for each region and subregion of countries
region_rows = {
    "AFRICA": {
        "EASTERN AFRICA": [9, 28],
        "MIDDLE AFRICA": [30, 38],
        "NORTHERN AFRICA": [40, 46],
        "SOUTHERN AFRICA": [48, 52],
        "WESTERN AFRICA": [54, 70]
    },
    "ASIA": {
        "CENTRAL ASIA": [73, 77],
        "EASTERN ASIA": [79, 85],
        "SOUTH-EASTERN ASIA": [87, 97],
        "SOUTHERN ASIA": [99, 107],
        "WESTERN ASIA": [109, 126]
    },
    "EUROPE": {
        "EASTERN EUROPE": [129, 138],
        "NORTHERN EUROPE": [140, 152],
        "SOUTHERN EUROPE": [154, 169],
        "WESTERN EUROPE": [171, 179]
    },
    "LATIN AMERICA AND THE CARIBBEAN": {
        "CARIBBEAN": [182, 207],
        "CENTRAL AMERICA": [209, 216],
        "SOUTH AMERICA": [218, 231],
        "NORTHERN AMERICA": [233, 237]
    },
    "OCEANIA": {
        "AUSTRALIA AND NEW ZEALAND": [240, 241],
        "MELANESIA": [243, 247],
        "MICRONESIA": [249, 255],
        "POLYNESIA": [257, 265]
    }
}

# The offset between the "sort" row in the table, and the actual row in the table.
row_offset = 16

# A list of the region and subregion names, so they can be conveniently skipped when parsing the table
forbidden = [r for r in region_rows]
for r in region_rows:
    for rr in region_rows[r]:
        forbidden.append(rr)


class MigrationGraph(object):
    def __init__(self):
        # Load the dataset for migrant flow
        migrants_flow = load_workbook(os.path.join(path, "data", "UN_MigrantStockByOriginAndDestination_2015.xlsx"))

        # Establish references to the correct table for each year of male migration
        male_migration = {
            "1990": migrants_flow["Table 2"],
            "1995": migrants_flow["Table 5"],
            "2000": migrants_flow["Table 8"],
            "2005": migrants_flow["Table 11"],
            "2010": migrants_flow["Table 14"],
            "2015": migrants_flow["Table 17"]
        }
        # Establish references to the correct table for each year of female migration
        female_migration = {
            "1990": migrants_flow["Table 3"],
            "1995": migrants_flow["Table 6"],
            "2000": migrants_flow["Table 9"],
            "2005": migrants_flow["Table 12"],
            "2010": migrants_flow["Table 15"],
            "2015": migrants_flow["Table 18"]
        }

        # A dictionary to hold every node, indexed by country name.
        self.nodes = {}

        # Initialize Nodes
        self.pop_nodes()

        # Initialize Link Structure
        for year in male_migration:
            self.pop_links(male_migration[year], year, "m")
        for year in female_migration:
            self.pop_links(female_migration[year], year, "f")

        # Perform a bunch of batch statistics on the graph
        self.totals()

    def pop_links(self, table, yr, gender):
        yr = str(yr)
        for col in table.iter_cols(min_col=origins[0], max_col=origins[1], min_row=25, max_row=281):
            source = table[col[0].column + "16"].value
            source_total = table[col[0].column + "17"].value
            self.nodes[source]['data']['emmigrant_pop'][yr][gender] = source_total
            for cell in col:
                target = table[country_names + str(cell.row)].value
                if target.upper() not in forbidden and cell.value is not None and int(cell.value) != 0:
                    adj = self.nodes[source]['out_neighbors']
                    tgt = self.nodes[target]['in_neighbors']
                    if target not in adj:
                        adj[target] = {}
                    if yr not in adj[target]:
                        adj[target][yr] = {}
                    if source not in tgt:
                        tgt[source] = {}
                    if yr not in tgt[source]:
                        tgt[source][yr] = {}
                    adj[target][yr][gender] = int(cell.value)
                    tgt[source][yr][gender] = int(cell.value)

    def pop_nodes(self):
        migrants_total = load_workbook(os.path.join(path, "data", "UN_MigrantStockTotal_2015.xlsx"))['Table 1']
        for region in region_rows:
            for subregion in region_rows[region]:
                rows = region_rows[region][subregion]
                for row in range(rows[0]+row_offset, rows[1]+1+row_offset):
                    source = migrants_total[country_names+str(row)].value
                    self.nodes[source] = {"in_neighbors": {}, "out_neighbors": {}, "data": {'immigrant_pop': {}, 'emmigrant_pop':{}}}
                    for year in total_pops:
                        self.nodes[source]["data"]['immigrant_pop'][year] = {}
                        self.nodes[source]["data"]['emmigrant_pop'][year] = {}
                        for category in total_pops[year]:
                            col = total_pops[year][category]
                            val = migrants_total[col+str(row)].value
                            self.nodes[source]["data"]['immigrant_pop'][year][category] = val

    def totals(self):
        for n in self.nodes:
            ep = self.nodes[n]['data']['emmigrant_pop']
            total = 0
            for yr in ep:
                ep[yr]['total'] = ep[yr]['m']+ep[yr]['f']
                total += ep[yr]['total']
            ep['total'] = total
            nn = self.nodes[n]['out_neighbors']
            max_val = 0
            for n2 in nn:
                c_total = 0
                for yr in nn[n2]:
                    nn[n2][yr]['total'] = 0
                    if 'm' in nn[n2][yr]:
                        nn[n2][yr]['total'] += nn[n2][yr]['m']
                    if 'f' in nn[n2][yr]:
                        nn[n2][yr]['total'] += nn[n2][yr]['f']
                    c_total += nn[n2][yr]['total']
                nn[n2]['total'] = c_total
                if c_total > max_val:
                    max_val = c_total
            self.nodes[n]['out_neighbors']['max_total'] = max_val


g = MigrationGraph()
with open(os.path.join(path, "data", "MigrationGraph.json"), "w") as f:
    json.dump(g.nodes, f)
